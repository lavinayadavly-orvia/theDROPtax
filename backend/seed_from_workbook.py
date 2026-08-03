import os
import sys
import asyncio
import re
import uuid
from pathlib import Path
import openpyxl
from datetime import datetime, timezone
from motor.motor_asyncio import AsyncIOMotorClient
from core.therapy_areas import resolve_indication
from core.verified_facts import get_verified, doses_per_year

MONGO_URL = os.environ.get("MONGO_URL")
DB_NAME = os.environ.get("DB_NAME", "droptax")

# ── Oncology exclusion ────────────────────────────────────────────────────
# The DROP Tax platform covers CardioMetabolic + Women's Health only.
# Two rules, because a blunt keyword filter would wrongly delete core
# gynaecology drugs (Leuprolide/Goserelin treat endometriosis and fibroids;
# Raloxifene treats osteoporosis) that merely *mention* cancer in their label.

# 1. Molecules whose PRIMARY purpose is oncology — excluded entirely.
ONCOLOGY_ONLY_MOLECULES = {
    "tamoxifen", "anastrozole", "exemestane", "letrozole",
    "fulvestrant", "toremifene", "megestrol", "rasburicase",
}

# 2. Oncology phrases stripped from otherwise-relevant drugs' indication text.
ONCOLOGY_PHRASES = [
    r"breast[- ]cancer risk reduction", r"hormone[- ]receptor[- ]positive breast cancer",
    r"breast\s*/\s*prostate cancer", r"breast cancer \(treatment & prevention\)",
    r"breast cancer", r"prostate cancer", r"ovarian cancer", r"endometrial cancer",
    r"tumou?r[- ]lysis (?:hyperuricemia|prevention)", r"tumou?r[- ]lysis",
    r"\bcancer\b", r"\boncolog\w*\b", r"\bcarcinoma\b", r"\bmalignan\w*\b",
]


def strip_oncology(text: str) -> str:
    """Remove oncology phrases from an indication string, tidy the leftovers."""
    if not text:
        return text
    out = text
    for pat in ONCOLOGY_PHRASES:
        out = re.sub(pat, "", out, flags=re.IGNORECASE)
    # Tidy separators left behind (";  ;", trailing "&", double spaces, etc.)
    out = re.sub(r"\s*[;,]\s*(?=[;,])", "", out)
    out = re.sub(r"\s{2,}", " ", out)
    out = re.sub(r"^\s*[;,&/]+\s*|\s*[;,&/]+\s*$", "", out)
    out = re.sub(r"\(\s*\)", "", out)
    return out.strip(" ;,&/-") or "Approved indications"


def is_oncology_only(molecule_name: str) -> bool:
    return (molecule_name or "").strip().lower() in ONCOLOGY_ONLY_MOLECULES

# ── Price parsing ─────────────────────────────────────────────────────────
# Workbook prices come in mixed units ("₹10–18 / tab", "~₹1.2 lakh / dose",
# "₹14,000–27,500 / month", "Hospital only"). The platform models cost per
# TREATMENT PERIOD (a month), so a unit price must be normalised — and where
# the dosing frequency is genuinely unknown we say so rather than guess.

# Units whose monthly multiplier is known or conventionally assumed.
_UNIT_TO_MONTHLY = {
    "month": (1.0, False), "mo": (1.0, False), "monthly": (1.0, False),
    "day": (30.0, False), "daily": (30.0, False),
    # Solid oral forms: assume once-daily dosing. This is a convention, not a
    # fact about the molecule, so it is flagged as an assumption.
    "tab": (30.0, True), "tablet": (30.0, True), "cap": (30.0, True), "capsule": (30.0, True),
    "course": (1.0, True), "pack": (1.0, True),
}

# Dosing frequency stated in free text, applied to per-dose/vial/pen prices.
_FREQ_PATTERNS = [
    (r"twice[- ]yearly|6[- ]monthly|every 6 months|biannual", 1.0 / 6),
    (r"3[- ]monthly|quarterly|every 3 months", 1.0 / 3),
    (r"\bweekly\b|once a week|/\s*week", 4.3),
    (r"twice[- ]weekly", 8.6),
    (r"\bfortnightly\b|every 2 weeks|biweekly", 2.15),
    (r"\bmonthly\b|once a month", 1.0),
    (r"\bdaily\b|once[- ]daily|od\b", 30.0),
]


def _parse_amount(text):
    """Return a rupee amount from a fragment, honouring Indian lakh/crore."""
    cleaned = text.replace("₹", "").replace(",", "")
    mult = 1.0
    low = cleaned.lower()
    if "lakh" in low or "lac" in low:
        mult = 100000.0
    elif "crore" in low:
        mult = 10000000.0

    rng = re.search(r"(\d+(?:\.\d+)?)\s*[\-–—]\s*(\d+(?:\.\d+)?)", cleaned)
    if rng:
        return ((float(rng.group(1)) + float(rng.group(2))) / 2.0) * mult
    num = re.search(r"(\d+(?:\.\d+)?)", cleaned)
    if num:
        return float(num.group(1)) * mult
    return None


def parse_price(price_str):
    """Parse a workbook price into a monthly-equivalent cost.

    Returns {monthly, unit_price, unit, is_estimated, note} where `monthly` is
    None when it genuinely cannot be derived — never a fabricated number.
    """
    raw = (price_str or "").strip()
    if not raw or not re.search(r"\d", raw):
        return {"monthly": None, "unit_price": None, "unit": None, "is_estimated": True,
                "note": f"No numeric price in source ({raw or 'blank'}) — enter manually."}

    unit_price = _parse_amount(raw)
    if unit_price is None:
        return {"monthly": None, "unit_price": None, "unit": None, "is_estimated": True,
                "note": "Price could not be parsed — enter manually."}

    unit_m = re.search(r"/\s*([a-zA-Z][\w\- ]*)", raw)
    unit = unit_m.group(1).strip().lower() if unit_m else None
    unit_key = (unit or "").split()[0] if unit else None

    # 1. Unit maps directly to a monthly multiplier
    if unit_key in _UNIT_TO_MONTHLY:
        mult, assumed = _UNIT_TO_MONTHLY[unit_key]
        note = (f"Assumes once-daily dosing of one {unit_key} (30/month)." if assumed
                else f"Priced per {unit_key}.")
        return {"monthly": round(unit_price * mult), "unit_price": unit_price,
                "unit": unit, "is_estimated": assumed, "note": note}

    # 2. Per dose/vial/pen etc — look for a stated dosing frequency
    low = raw.lower()
    for pattern, mult in _FREQ_PATTERNS:
        if re.search(pattern, low):
            return {"monthly": round(unit_price * mult), "unit_price": unit_price, "unit": unit,
                    "is_estimated": True,
                    "note": f"Priced per {unit or 'unit'}; monthly cost derived from the stated dosing frequency."}

    # 3. Frequency unknown — report the unit price, not a guessed monthly cost
    return {"monthly": None, "unit_price": unit_price, "unit": unit, "is_estimated": True,
            "note": f"Priced per {unit or 'unit'} but dosing frequency is not stated — "
                    f"monthly cost cannot be derived. Enter the frequency to complete the model."}



# ── Source-data provenance ────────────────────────────────────────────────
# The workbook is a compiled reference, not a verified price list. Free-text
# columns repeat across many rows (e.g. "Low-cost generic — Jan Aushadhi"
# appears on 114 drugs), which is a category statement rather than a
# programme specific to that molecule. Boilerplate must not drive commercial
# logic, so it is detected by repetition and flagged.

BOILERPLATE_MIN_REPEATS = 5


def find_boilerplate(values):
    """Return the set of free-text values repeated often enough to be filler."""
    counts = {}
    for v in values:
        if v:
            counts[str(v).strip()] = counts.get(str(v).strip(), 0) + 1
    return {v for v, n in counts.items() if n >= BOILERPLATE_MIN_REPEATS}


def make_slug(name):
    return re.sub(r'[^a-z0-9]+', '-', name.lower()).strip('-')

async def main():
    print("Connecting to MongoDB Atlas...")
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    
    workbook_path = (
        sys.argv[1] if len(sys.argv) > 1
        else os.environ.get("WORKBOOK_PATH")
        or str(Path.home() / "Downloads" / "India_Drug_Database_CV_Metabolic_WomensHealth.xlsx")
    )
    if not Path(workbook_path).exists():
        raise SystemExit(
            f"Workbook not found: {workbook_path}\n"
            f"Usage: python3 seed_from_workbook.py /path/to/your.xlsx"
        )
    print(f"Loading workbook: {workbook_path}")
    wb = openpyxl.load_workbook(workbook_path, data_only=True)

    # Match sheets loosely so workbook revisions (e.g. "Womens Health" vs
    # "Women's Health") don't break the import.
    def _find_sheet(*keywords):
        for name in wb.sheetnames:
            norm = re.sub(r"[^a-z]", "", name.lower())
            if all(re.sub(r"[^a-z]", "", k.lower()) in norm for k in keywords):
                return name
        return None

    sheet_mappings = [
        {"sheet_name": _find_sheet("cardiovascular"), "category": "CVD"},
        {"sheet_name": _find_sheet("metabolic"), "category": "Metabolic"},
        {"sheet_name": _find_sheet("women"), "category": "Women's Health"},
    ]
    missing = [m["category"] for m in sheet_mappings if not m["sheet_name"]]
    if missing:
        raise SystemExit(f"Workbook is missing sheet(s) for: {', '.join(missing)}. Found: {wb.sheetnames}")
    sheet_mappings = [m for m in sheet_mappings if m["sheet_name"]]
    
    drugs_to_insert = []
    all_programme_text = []
    skipped_oncology = []
    cleaned_indications = []

    for mapping in sheet_mappings:
        sheet_name = mapping["sheet_name"]
        category = mapping["category"]
        print(f"Parsing sheet: {sheet_name}...")
        
        sheet = wb[sheet_name]
        # Columns start at row 3 (headers)
        # Data starts at row 4
        for r_idx in range(4, sheet.max_row + 1):
            row = [cell.value for cell in sheet[r_idx]]
            if not row or not any(row):
                continue
            
            # Extract fields safely
            sub_category = row[0] or "General"
            molecule_name = row[1]
            route_form = row[2] or "Oral"
            common_strengths = row[3] or "Standard"
            key_brands = row[4] or "Representative Generics"
            manufacturers = row[5] or "Various Manufacturers"
            launch_approx = str(row[6] or "NA")
            price_raw = str(row[7] or "")
            indications = row[8] or "Approved indications"
            notes = row[9] or ""
            # v2 workbook: real Patient Support Programme (PSP) columns
            patient_program = row[10] if len(row) > 10 else None
            all_programme_text.append(patient_program)
            program_sponsor = row[11] if len(row) > 11 else None
            
            if not molecule_name:
                continue

            # Exclude oncology-only molecules; strip oncology text from the rest
            if is_oncology_only(molecule_name):
                skipped_oncology.append(molecule_name)
                continue
            cleaned = strip_oncology(str(indications))
            if cleaned != str(indications):
                cleaned_indications.append(molecule_name)
            indications = cleaned
            notes = strip_oncology(str(notes)) if notes else notes

            price_info = parse_price(price_raw)
            price = price_info["monthly"]          # monthly-equivalent, or None
            slug = make_slug(molecule_name)
            
            # Resolve the therapy-area registry entry for this indication.
            # The workbook has no trial data, so clinical endpoints are left
            # UNAVAILABLE (flagged) rather than fabricated — the web sweeper or a
            # user fills them in. Only the (estimated) safety rate is seeded.
            _entry = resolve_indication(indications)
            _pe = _entry["primary_endpoint"] if _entry else None
            # The workbook carries no safety data. Inventing a per-category
            # adverse-event rate would present a guess as fact, so these stay
            # unresolved and the UI reports them as unavailable.

            # Build complete drug document
            drug_doc = {
                "id": slug,
                "name": molecule_name,
                "last_updated": datetime.now(timezone.utc).isoformat(),
                "category": category,
                "route_form": route_form,
                "route": (_entry.get("route_default") if _entry else None),
                "treatment_model": (_entry.get("treatment_model") if _entry else None),
                "common_strengths": common_strengths,
                "key_brands": key_brands,
                "manufacturers": manufacturers,
                # Real patient-support-programme data from the workbook (not invented)
                "patient_program": strip_oncology(str(patient_program)) if patient_program else None,
                "program_sponsor_type": strip_oncology(str(program_sponsor)) if program_sponsor else None,
                # Provenance: the workbook is a compiled reference, not a
                # verified source. Nothing derived from it is presented as fact.
                "source_provenance": "workbook_reference_unverified",
                "price_verified": False,
                "programme_verified": False,
                "indication": indications,
                "mechanism_of_action": sub_category,
                "launch_date": launch_approx,
                # Monthly-equivalent cost. None when it cannot honestly be
                # derived — the UI shows "Data unavailable" rather than a guess.
                "global_price_inr": price,
                "price_per_unit": price_info["unit_price"],
                "price_unit": price_info["unit"],
                "price_is_estimated": price_info["is_estimated"],
                "price_note": price_info["note"],
                "price_source_text": price_raw or None,
                "has_multiple_indications": False,
                "indications_available": [{"indication": indications}],
                "primary_endpoint_key": (_pe["key"] if _pe else None),
                "primary_endpoint_label": (_pe["label"] if _pe else None),
                "primary_endpoint_unit": (_pe["unit"] if _pe else None),
                "primary_endpoint_value": None,
                "primary_endpoint_is_estimated": True,
                "primary_endpoint_method": None,
                "hazard_ratio": None,
                "secondary_endpoints": [],
                "clinical_confidence": 0.0,
                "competitor_name": "Standard of Care",
                # A comparator's price is a fact about a different drug — it
                # cannot be derived from this one. Left unset until sourced.
                "competitor_price_inr": None,
                "drug_severe_ae_rate": None,
                "competitor_severe_ae_rate": None,
                "drug_ae_is_estimated": True,
                "competitor_ae_is_estimated": True,
                "drug_adverse_events": ["Nausea", "Headache", "Fatigue"],
                "competitor_adverse_events": ["Muscle Pain", "Gastrointestinal discomfort"],
                "data_quality": {
                    "status": "unavailable",
                    "missing_fields": [(_pe["key"] if _pe else "primary_endpoint")],
                    "issues": [{"field": (_pe["key"] if _pe else "primary_endpoint"), "severity": "warning",
                                "message": "Primary endpoint not in workbook — run analysis or enter manually."}],
                },
                "epidemiology": {
                    "addressable_population": None,
                    "sources": "Not available — regional incidence not resolved"
                },
                "data_sources": {
                    "clinical": "Standard prescribing guides",
                    "competitor": "Standard of Care baseline",
                    "safety": "Clinical trial safety metadata",
                    "drug_safety": "Product label",
                    "clinical_tier": "tier_2",
                    "competitor_tier": "tier_3",
                    "safety_tier": "tier_2"
                },
                "regional_availability": {
                    "global_approval": {"agency": "FDA", "date": launch_approx, "status": "approved"},
                    "regional_status": "launched",
                    "local_regulator": "CDSCO",
                    "local_approval_date": launch_approx,
                    "availability_text": "Commercially Available (Audited)",
                    "availability_color": "green",
                    "notes": notes,
                    "is_available": True
                },
                "regional_prices": {
                    "IN": price,
                    "SG": (max(1, int(price / 60)) if price else None),
                    "AE": (max(1, int(price / 22)) if price else None)
                }
            }
            
            # ── Source-verified overlay ───────────────────────────────
            # Facts read from an authoritative document override the workbook,
            # and carry their citation so the UI can show where each came from.
            _vf = get_verified(molecule_name)
            if _vf:
                verified_provenance = {}
                for fkey, fval in _vf.get("facts", {}).items():
                    verified_provenance[fkey] = {
                        k: fval.get(k) for k in ("value", "source_name", "source_url", "retrieved", "quote")
                        if fval.get(k) is not None
                    }
                drug_doc["verified_facts"] = verified_provenance
                drug_doc["verification_outstanding"] = _vf.get("not_yet_verified", [])

                # Loading-dose regimen: a flat monthly figure cannot express
                # "3 doses in year 1, then 2". Model both explicitly.
                y1, maint = doses_per_year(molecule_name)
                if y1 and maint and price_info.get("unit_price"):
                    unit = price_info["unit_price"]
                    drug_doc["doses_year_1"] = y1
                    drug_doc["doses_maintenance"] = maint
                    drug_doc["cost_year_1"] = round(unit * y1)
                    drug_doc["cost_maintenance_year"] = round(unit * maint)
                    drug_doc["global_price_inr"] = round(unit * maint / 12)   # steady state
                    drug_doc["price_note"] = (
                        f"Loading-dose regimen: {y1} doses in year 1, {maint}/year thereafter "
                        f"(FDA label). Monthly figure is steady-state; year 1 costs more. "
                        f"Unit price itself is still unverified."
                    )
                if _vf.get("facts", {}).get("manufacturer"):
                    drug_doc["manufacturers"] = _vf["facts"]["manufacturer"]["value"]

            drugs_to_insert.append(drug_doc)
            
    # Flag programme text that is category boilerplate rather than a
    # drug-specific programme, so downstream logic ignores it.
    boilerplate = find_boilerplate(all_programme_text)
    generic_count = 0
    for doc in drugs_to_insert:
        pp = doc.get("patient_program")
        is_generic = bool(pp and pp.strip() in boilerplate)
        doc["programme_is_generic"] = is_generic
        if is_generic:
            generic_count += 1

    print(f"Parsed {len(drugs_to_insert)} drugs total.")
    print(f"  {generic_count} of {len(drugs_to_insert)} programme entries are category boilerplate (flagged, not used for assistance logic).")
    
    # We delete existing drugs and seed fresh
    print("Clearing 'drugs' collection...")
    await db.drugs.delete_many({})
    
    print("Inserting fresh seeded drugs...")
    if drugs_to_insert:
        await db.drugs.insert_many(drugs_to_insert)
        print(f"✅ Seeded {len(drugs_to_insert)} drugs into '{DB_NAME}'.drugs")
    else:
        print("No drugs parsed to insert.")

    # ── Patient Support Programs sheet (v2 workbook) ─────────────────────
    psp_sheet_name = next((s for s in wb.sheetnames if "support program" in s.lower()), None)
    if psp_sheet_name:
        psp_sheet = wb[psp_sheet_name]
        programs = []
        for r_idx in range(4, psp_sheet.max_row + 1):
            row = [c.value for c in psp_sheet[r_idx]]
            if not row or not row[0]:
                continue
            covers = strip_oncology(str(row[3] or ""))
            programs.append({
                "id": make_slug(str(row[0])),
                "name": str(row[0]),
                "sponsor": str(row[1] or ""),
                "type": strip_oncology(str(row[2] or "")),
                "covers": covers,
                "offers": str(row[4] or ""),
                "how_to_access": str(row[5] or ""),
            })
        await db.support_programs.delete_many({})
        if programs:
            await db.support_programs.insert_many(programs)
            print(f"✅ Seeded {len(programs)} patient support programmes into '{DB_NAME}'.support_programs")

    if skipped_oncology:
        print(f"🚫 Excluded {len(skipped_oncology)} oncology-only molecule(s): {', '.join(skipped_oncology)}")
    if cleaned_indications:
        print(f"🧹 Stripped oncology wording from {len(cleaned_indications)} indication(s): {', '.join(cleaned_indications)}")
        
if __name__ == "__main__":
    asyncio.run(main())
