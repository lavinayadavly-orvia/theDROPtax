"""
Read a completed review worksheet back into the platform.

A row confirmed by a human becomes a verified price, stored with the reviewer's
name, the date and the listing URL — the same provenance discipline applied to
FDA label facts. A row rejected by a human is recorded as rejected, so the
automated pipeline cannot quietly reinstate it on the next run.

Usage
-----
    python3 ingest_review.py ~/Downloads/DROP_Tax_Review_20260803.xlsx --reviewer "Name"
    python3 ingest_review.py <file> --dry-run
"""
import os
import sys
import json
import argparse
import asyncio
from datetime import date

import openpyxl
from motor.motor_asyncio import AsyncIOMotorClient

MONGO_URL = os.environ.get("MONGO_URL")
DB_NAME = os.environ.get("DB_NAME", "droptax")


def read_verdicts(path):
    wb = openpyxl.load_workbook(path)
    ws = wb["To Review"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    def cell(vals, name):
        i = idx.get(name)
        return vals[i] if i is not None and i < len(vals) else None

    verdicts = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        product_ok = str(cell(vals, "Correct product?") or "").strip().lower()
        price_ok = str(cell(vals, "Correct price?") or "").strip().lower()
        if not product_ok and not price_ok:
            continue                                    # not reviewed yet
        link = ws.cell(r, (idx.get("Open listing") or 9) + 1)
        verdicts.append({
            "molecule": cell(vals, "Molecule"),
            "retailer": cell(vals, "Retailer"),
            "listing": cell(vals, "Listing says"),
            "strength": cell(vals, "Strength"),
            "pack": cell(vals, "Pack"),
            "price": cell(vals, "Selling price (INR)"),
            "mrp": cell(vals, "MRP (INR)"),
            "product_ok": product_ok,
            "price_ok": price_ok,
            "notes": cell(vals, "Notes"),
            "source_url": (link.hyperlink.target if link.hyperlink else None),
        })
    return verdicts


async def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("worksheet")
    ap.add_argument("--reviewer", default=os.environ.get("USER", "unknown"))
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    if not os.path.exists(args.worksheet):
        raise SystemExit(f"Worksheet not found: {args.worksheet}")
    verdicts = read_verdicts(args.worksheet)
    if not verdicts:
        raise SystemExit("No rows have been reviewed yet — fill in the "
                         "'Correct product?' / 'Correct price?' columns first.")

    confirmed = [v for v in verdicts if v["product_ok"] == "yes" and v["price_ok"] == "yes"]
    rejected = [v for v in verdicts if "no" in (v["product_ok"], v["price_ok"])]
    unsure = [v for v in verdicts if v not in confirmed and v not in rejected]

    print(f"Reviewed rows: {len(verdicts)}")
    print(f"  confirmed : {len(confirmed)}")
    print(f"  rejected  : {len(rejected)}")
    print(f"  unsure    : {len(unsure)}")

    if args.dry_run:
        for v in confirmed[:10]:
            print(f"    would verify {v['molecule']} @ {v['price']} ({v['retailer']})")
        return

    if not MONGO_URL:
        raise SystemExit("MONGO_URL is not set.")
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    today = date.today().isoformat()

    n_ok = n_rej = 0
    for v in confirmed:
        # A human-confirmed listing is a verified price, with attribution.
        await db.drugs.update_one(
            {"name": v["molecule"]},
            {"$set": {
                "verified_price": {
                    "selling_price_inr": v["price"],
                    "mrp_inr": v["mrp"],
                    "strength": v["strength"],
                    "pack": v["pack"],
                    "retailer": v["retailer"],
                    "listing": v["listing"],
                    "source_url": v["source_url"],
                    "confirmed_by": args.reviewer,
                    "confirmed_on": today,
                    "note": v["notes"],
                },
                "price_verified": True,
            }})
        n_ok += 1

    for v in rejected:
        # Recorded so the pipeline cannot silently reinstate a rejected match.
        await db.price_rejections.update_one(
            {"molecule": v["molecule"], "source_url": v["source_url"]},
            {"$set": {
                "molecule": v["molecule"], "retailer": v["retailer"],
                "listing": v["listing"], "price": v["price"],
                "source_url": v["source_url"], "reason": v["notes"],
                "rejected_by": args.reviewer, "rejected_on": today,
            }}, upsert=True)
        n_rej += 1

    print(f"\n✅ {n_ok} price(s) marked verified against {args.reviewer}")
    print(f"   {n_rej} rejection(s) recorded so they are not reinstated")
    client.close()


if __name__ == "__main__":
    asyncio.run(main())
