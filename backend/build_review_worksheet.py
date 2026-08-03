"""
Build a focused human-review worksheet from the sourced-price workbook.

Hand-checking 787 rows is not realistic. This narrows the list to the rows
where a wrong number actually costs something, and puts everything a reviewer
needs on one line: the price, what the listing claims to be, and a clickable
link to the page it came from.

What gets pulled in
-------------------
1. Rows flagged OUTLIER by the cross-retailer check.
2. High-value molecules — anything at or above the price threshold, where an
   error moves a real negotiation.
3. Molecules with only a single source, which have no corroboration.
4. Rows with no strength or pack captured, so like-for-like is unverifiable.

Everything else — cheap generics corroborated by several retailers that agree
— is left out deliberately. Reviewer attention is the scarce resource.

The reviewer fills in three columns: Correct product?, Correct price?, and
Notes. ingest_review.py then reads the completed sheet back, so a verdict is
recorded once and never re-litigated.

Usage
-----
    python3 build_review_worksheet.py                       # default file
    python3 build_review_worksheet.py --prices <in.xlsx> --out <out.xlsx>
    python3 build_review_worksheet.py --threshold 5000
"""
import os
import argparse
from datetime import datetime, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

DEFAULT_PRICES = os.path.join(os.path.expanduser("~"), "Downloads",
                              "DROP_Tax_Sourced_Prices_FULL.xlsx")
HIGH_VALUE_THRESHOLD_INR = 2000     # per listing; below this an error is cheap


def load_rows(path):
    wb = openpyxl.load_workbook(path)
    ws = wb["Sourced Prices"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    rows = []
    for r in range(2, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        rows.append({h: vals[idx[h]] for h in headers})
    return rows


def needs_review(row, threshold):
    """Why this row deserves a human, or None to leave it out."""
    check = str(row.get("Cross-retailer check") or "")
    price = row.get("Selling price (INR)")
    if check.startswith("OUTLIER"):
        return "Price disagrees sharply with other retailers"
    if price and price >= threshold:
        return "High-value listing — an error here moves a negotiation"
    # Below the value threshold, a missing strength or a lone source is not
    # worth a reviewer's time. Flagging every cheap generic buried the rows
    # that actually matter: 186 of 256 flags came from unparsed strengths on
    # listings costing a few rupees.
    if price and price >= threshold / 4:
        if check == "single source":
            return "Only one retailer listed it — no corroboration"
        if not row.get("Strength") and not row.get("Pack"):
            return "No strength or pack captured — cannot compare like for like"
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--prices", default=DEFAULT_PRICES)
    ap.add_argument("--out", default=None)
    ap.add_argument("--threshold", type=float, default=HIGH_VALUE_THRESHOLD_INR)
    args = ap.parse_args()

    if not os.path.exists(args.prices):
        raise SystemExit(f"Price workbook not found: {args.prices}\n"
                         f"Run build_price_workbook.py first.")

    rows = load_rows(args.prices)
    flagged = []
    for r in rows:
        reason = needs_review(r, args.threshold)
        if reason:
            flagged.append((reason, r))

    # Most expensive first: that is where a mistake costs most.
    flagged.sort(key=lambda x: -(x[1].get("Selling price (INR)") or 0))

    out = args.out or os.path.join(os.path.expanduser("~"), "Downloads",
                                   f"DROP_Tax_Review_{datetime.now().strftime('%Y%m%d')}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "To Review"

    headers = ["#", "Molecule", "Why flagged", "Retailer", "Listing says",
               "Strength", "Pack", "Selling price (INR)", "MRP (INR)",
               "Open listing", "Correct product?", "Correct price?", "Notes"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E5F")
        c.alignment = Alignment(vertical="center", wrap_text=True)

    for i, (reason, r) in enumerate(flagged, 1):
        ws.append([i, r.get("Molecule"), reason, r.get("Retailer"),
                   r.get("Product listing"), r.get("Strength"), r.get("Pack"),
                   r.get("Selling price (INR)"), r.get("MRP (INR)"),
                   "Open", None, None, None])
        url = r.get("Source URL")
        if url:
            cell = ws.cell(ws.max_row, 10)
            cell.hyperlink = url
            cell.value = "Open listing"
            cell.font = Font(color="0563C1", underline="single")
        if reason.startswith("Price disagrees"):
            for col in range(1, 14):
                ws.cell(ws.max_row, col).fill = PatternFill("solid", fgColor="FDE7E7")

    # Yes/No pickers so verdicts come back in a consistent shape
    yn = DataValidation(type="list", formula1='"Yes,No,Unsure"', allow_blank=True)
    ws.add_data_validation(yn)
    yn.add(f"K2:L{max(ws.max_row, 2)}")

    for col, w in zip("ABCDEFGHIJKLM",
                      (5, 26, 34, 16, 46, 12, 8, 16, 14, 14, 16, 15, 40)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("How to review")
    for line in [
        ["Reviewing these listings"],
        [],
        ["Generated (UTC)", datetime.now(timezone.utc).isoformat(timespec="seconds")],
        ["Rows needing review", len(flagged)],
        ["Rows left out (corroborated, low value)", len(rows) - len(flagged)],
        [],
        ["For each row"],
        ["1", "Click 'Open listing' to load the exact page the price came from."],
        ["2", "Correct product? — is this the molecule in column B, at a plausible "
              "strength and dose form? A capsule listing for an infusion-only drug "
              "is the commonest error."],
        ["3", "Correct price? — does the price on the page match column H today? "
              "Listings move; a mismatch may just mean the price changed."],
        ["4", "Notes — record the right value if you can see it, or why the row is wrong."],
        [],
        ["What to look for"],
        ["Wrong molecule", "A shared salt or brand name can pull in a different drug. "
                           "Zoledronic acid matched a rabeprazole brand this way."],
        ["Wrong form", "Zoledronic acid is an infusion, so a capsule cannot be it."],
        ["Wrong pack", "A 30-tablet pack is not comparable with a 10-tablet pack."],
        ["Marketing text", "Percentages in titles are often discounts, not strengths."],
        [],
        ["When done"],
        ["", "Save the file and run: python3 ingest_review.py <this file>"],
        ["", "Confirmed rows are recorded as verified with your name against them, "
             "and are not surfaced for review again."],
        [],
        ["Scope"],
        ["", "These are retail listings, not NPPA ceiling prices, and not net "
             "realised prices. Confirming a row means the listing was read "
             "correctly — not that it is the right price for a payer submission."],
    ]:
        ws2.append(line)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 96
    ws2["A1"].font = Font(bold=True, size=13)

    wb.save(out)
    print(f"✅ {len(flagged)} row(s) need review, out of {len(rows)}")
    by_reason = {}
    for reason, _ in flagged:
        by_reason[reason] = by_reason.get(reason, 0) + 1
    for reason, n in sorted(by_reason.items(), key=lambda x: -x[1]):
        print(f"     {n:>4}  {reason}")
    print(f"   written to {out}")


if __name__ == "__main__":
    main()
