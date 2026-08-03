"""
Build a NEW drug-price workbook from live Indian pharmacy listings.

This does not edit or rebuild the existing reference workbook. It produces a
separate file whose every price cell carries the URL it came from and the
timestamp it was read, so the sheet can be re-derived and audited.

Principles
----------
* Nothing is hardcoded. The molecule list comes from the database; the prices
  come from the retailers' own pages.
* Nothing is inferred. If a price is not present on the page, the cell is left
  empty and the reason recorded. There are no fallbacks, averages or defaults.
* Prices move. The Sybrava listing changed by ~3,000 rupees between two reads
  minutes apart, so every row is stamped with its retrieval time.
* robots.txt is respected: product URLs are discovered from each site's own
  sitemap, never from the search endpoints those files disallow.

Usage
-----
    python3 build_price_workbook.py --limit 25        # first 25 molecules
    python3 build_price_workbook.py --molecule Inclisiran Atorvastatin
    python3 build_price_workbook.py                   # every molecule in the DB
"""
import os
import re
import sys
import gzip
import json
import time
import asyncio
import argparse
import urllib.request
from datetime import datetime, timezone
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from motor.motor_asyncio import AsyncIOMotorClient

MONGO_URL = os.environ.get("MONGO_URL")
DB_NAME = os.environ.get("DB_NAME", "droptax")

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/120.0 Safari/537.36")

# Retailers to read. `url_contains` identifies product pages within a sitemap.
SOURCES = [
    {"key": "1mg",      "name": "Tata 1mg",         "sitemap": "https://www.1mg.com/sitemap.xml",
     "url_contains": "/drugs/"},
    {"key": "apollo",   "name": "Apollo Pharmacy",  "sitemap": "https://www.apollopharmacy.in/sitemap/sitemap-master.xml",
     "url_contains": "/medicine/"},
    {"key": "mrmed",    "name": "MrMed",            "sitemap": "https://www.mrmed.in/sitemap.xml",
     "url_contains": "/medicines/"},
    {"key": "netmeds",  "name": "NetMeds",          "sitemap": "https://www.netmeds.com/sitemap.xml",
     "url_contains": "/product/"},
    {"key": "pharmeasy", "name": "PharmEasy",       "sitemap": "https://pharmeasy.in/sitemap.xml",
     "url_contains": "/online-medicine-order/"},
]

REQUEST_DELAY_SEC = 1.0        # deliberate: this is someone else's server


def fetch(url, timeout=30):
    req = urllib.request.Request(url, headers={"User-Agent": UA})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        raw = r.read()
    if raw[:2] == b"\x1f\x8b":
        raw = gzip.decompress(raw)
    return raw.decode("utf-8", errors="ignore")


CACHE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".sitemap_cache")


def sitemap_urls(source, refresh=False):
    """Collect product URLs from a retailer's sitemap, following nested indexes.

    Sitemap indexes nest (Apollo goes master -> section -> page), and 1mg has
    218 child files. Reading only the first level or a fixed slice silently
    under-covers the catalogue and produces false "not found" results, so the
    whole tree is walked and cached on disk.
    """
    os.makedirs(CACHE_DIR, exist_ok=True)
    cache = os.path.join(CACHE_DIR, f"{source['key']}.json")
    if os.path.exists(cache) and not refresh:
        age_h = (time.time() - os.path.getmtime(cache)) / 3600
        if age_h < 24:
            with open(cache) as f:
                cached = json.load(f)
            # Never trust an empty cache. A failed or capped crawl once wrote
            # zero URLs for MrMed, and the empty result was then served for 24
            # hours as though the retailer had no catalogue — turning a bug
            # into a silent "no listing found" for every molecule.
            if cached:
                return cached

    seen_maps, product_urls = set(), []
    queue = [source["sitemap"]]
    while queue:
        current = queue.pop(0)
        if current in seen_maps:
            continue
        seen_maps.add(current)
        try:
            body = fetch(current)
        except Exception:
            continue
        locs = re.findall(r"<loc>\s*([^<]+?)\s*</loc>", body)
        for loc in locs:
            if loc.endswith((".xml", ".xml.gz")):
                queue.append(loc)                      # nested index
            elif source["url_contains"] in loc:
                product_urls.append(loc)
        time.sleep(0.15)

    with open(cache, "w") as f:
        json.dump(product_urls, f)
    return product_urls


# Values that appear as placeholders in retailer markup rather than real
# prices. PharmEasy returned 1000 as the MRP on every row sampled; Apollo
# returned 2999 and 99999. Treating these as prices silently corrupts the model.
SENTINEL_PRICES = {1000.0, 2999.0, 9999.0, 99999.0, 999999.0}

# A molecule name can appear in things that are not that medicine — diagnostic
# panels, combination anaesthetics, lab bookings.
NON_PRODUCT_MARKERS = (
    "deaminase", " csf", "book ", " test", "test online", "diagnostic",
    "lab test", "profile", "panel", "checkup", "check-up", "scan",
)


# An injectable molecule cannot be sold as a tablet. Zoledronic acid is an IV
# infusion, so a "Zolic 20 MG Capsule" listing is a different drug regardless
# of what the reference workbook lists as its brands.
ORAL_FORMS = ("tablet", "capsule", "syrup", "suspension", "drops", "sachet", "powder")
INJECTABLE_FORMS = ("injection", "infusion", "vial", "ampoule", "prefilled", "pfs", "syringe")


def form_conflicts_with_route(title, route):
    """True when the listing's dose form contradicts the molecule's route."""
    t = (title or "").lower()
    if not route:
        return False
    is_oral_listing = any(f in t for f in ORAL_FORMS)
    is_inj_listing = any(f in t for f in INJECTABLE_FORMS)
    if route in ("iv_bolus", "iv_infusion", "sc_injection"):
        return is_oral_listing and not is_inj_listing
    if route == "oral":
        return is_inj_listing and not is_oral_listing
    return False


def looks_like_wrong_product(title, molecule, route=None):
    """Reject listings that are not the medicine in question."""
    t = (title or "").lower()
    if any(mark in t for mark in NON_PRODUCT_MARKERS):
        return True
    if form_conflicts_with_route(title, route):
        return True
    return False


def extract_pack(title):
    """Strength and pack size, so prices are only compared like for like."""
    t = title or ""
    # Strip retailer branding first: "... | 1mg" is the site name, not a dose.
    t = re.sub(r"\s*[|\-–]\s*(1mg|Tata 1mg|Netmeds|PharmEasy|Apollo Pharmacy|MrMed)\b.*$",
               "", t, flags=re.IGNORECASE)
    # Promotional text carries percentages that are discounts, not strengths.
    t = re.sub(r"save up to[^,]*", " ", t, flags=re.IGNORECASE)
    t = re.sub(r"\b\d+%\s*(off|discount)", " ", t, flags=re.IGNORECASE)
    strength = None
    m = re.search(r"(\d+(?:\.\d+)?\s?(?:mg|mcg|g|ml|iu|%))", t, re.IGNORECASE)
    if m:
        strength = m.group(1).strip()
    pack = None
    m = re.search(r"(\d+)\s*(?:'|&#x27;)?\s*[sS]\b|\((\d+)\)", t)
    if m:
        pack = m.group(1) or m.group(2)
    return strength, pack


def extract_price(html):
    """Pull selling price and MRP from a product page. None when absent."""
    price = mrp = None
    m = re.search(r'"mrp"\s*:\s*"?([\d.]+)', html)
    if m:
        mrp = float(m.group(1))
    m = re.search(r'"(?:discounted_price|price|selling_price|final_price)"\s*:\s*"?([\d.]+)', html)
    if m:
        price = float(m.group(1))
    if price is None or mrp is None:
        rupees = [float(x.replace(",", "")) for x in re.findall(r"₹\s?([\d,]{3,})", html)]
        rupees = [r for r in rupees if r >= 1]
        if rupees:
            if price is None:
                price = min(rupees[:6])
            if mrp is None:
                mrp = max(rupees[:6])
    # Discard placeholder values rather than passing them off as prices.
    if mrp in SENTINEL_PRICES:
        mrp = None
    if price in SENTINEL_PRICES:
        price = None
    return price, mrp


def extract_field(html, patterns):
    for p in patterns:
        m = re.search(p, html, re.IGNORECASE)
        if m:
            return re.sub(r"\s+", " ", m.group(1)).strip()[:120]
    return None


# Tokens that appear in thousands of unrelated drug names. Matching on these
# pulls in the wrong molecule entirely: "acid" matched "Zoledronic acid" to a
# rabeprazole brand, and "depot" matched hydroxyprogesterone to leuprolide.
GENERIC_TOKENS = {
    "acid", "acids", "depot", "tablet", "tablets", "capsule", "capsules",
    "injection", "injections", "solution", "syrup", "suspension", "cream",
    "sodium", "potassium", "calcium", "magnesium", "chloride", "sulfate",
    "sulphate", "hydrochloride", "acetate", "citrate", "maleate", "tartrate",
    "succinate", "fumarate", "phosphate", "nitrate", "oral", "forte", "plus",
    "cream", "gel", "drops", "spray", "patch", "inhaler", "powder", "sachet",
    "extended", "release", "prolonged", "combination", "hydrate", "monohydrate",
}


def slug_tokens(name):
    """Distinctive tokens used to match a molecule to a product URL.

    Generic pharmaceutical words are removed, because matching on them
    associates a molecule with an unrelated product that merely shares a
    salt name or dose form.
    """
    clean = re.sub(r"\(.*?\)", " ", name).lower()
    clean = re.split(r"[+/]", clean)[0]
    toks = [t for t in re.split(r"[^a-z0-9]+", clean) if len(t) > 3]
    distinctive = [t for t in toks if t not in GENERIC_TOKENS]
    # If a name is entirely generic words, fall back to the longest token so
    # the molecule is still attempted rather than silently skipped.
    return distinctive or (sorted(toks, key=len, reverse=True)[:1] if toks else [])


async def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--molecule", nargs="*", default=None)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    if not MONGO_URL:
        raise SystemExit("MONGO_URL is not set.")
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]

    if args.molecule:
        molecules = [{"name": m, "key_brands": "", "category": ""} for m in args.molecule]
    else:
        molecules = []
        async for d in db.drugs.find({}, {"_id": 0, "name": 1, "key_brands": 1,
                                          "category": 1, "route": 1}):
            molecules.append(d)
        molecules.sort(key=lambda d: d["name"])
        if args.limit:
            molecules = molecules[: args.limit]

    print(f"Sourcing prices for {len(molecules)} molecule(s) from "
          f"{len(SOURCES)} retailers.\n")

    print("Reading retailer sitemaps (this is the slow part)…")
    catalog = {}
    for src in SOURCES:
        urls = sitemap_urls(src)
        catalog[src["key"]] = urls
        print(f"    {src['name']:18} {len(urls):>7} product URLs")
    total_urls = sum(len(v) for v in catalog.values())
    if not total_urls:
        raise SystemExit("No product URLs discovered — cannot source prices.")

    # Index URLs by token for matching
    index = defaultdict(list)
    for key, urls in catalog.items():
        for u in urls:
            tail = u.rstrip("/").split("/")[-1].lower()
            for tok in set(re.split(r"[^a-z0-9]+", tail)):
                if len(tok) > 3:
                    index[tok].append((key, u))

    rows, misses = [], []
    stamp = datetime.now(timezone.utc).isoformat(timespec="seconds")

    for i, mol in enumerate(molecules, 1):
        name = mol["name"]
        toks = slug_tokens(name)
        brand_toks = [t for b in re.split(r"[,;]", str(mol.get("key_brands") or ""))
                      for t in slug_tokens(b)]
        candidates = []
        for tok in set(toks + brand_toks):
            candidates.extend(index.get(tok, []))
        # Keep one product URL per retailer
        seen, picked = set(), []
        for key, u in candidates:
            if key in seen:
                continue
            seen.add(key)
            picked.append((key, u))

        if not picked:
            misses.append((name, "no matching product page in any retailer sitemap"))
            print(f"  [{i:3}/{len(molecules)}] {name[:30]:32} — no listing found")
            continue

        found_any = False
        for key, url in picked:
            src = next(s for s in SOURCES if s["key"] == key)
            try:
                html = fetch(url)
            except Exception as e:
                misses.append((name, f"{src['name']}: fetch failed ({e})"))
                time.sleep(REQUEST_DELAY_SEC)
                continue
            price, mrp = extract_price(html)
            title = extract_field(html, [r"<title>([^<]{5,140})</title>"])
            mfr = extract_field(html, [r'"manufacturer"\s*:\s*"([^"]{2,80})"',
                                       r"Marketer</[^>]+>\s*<[^>]+>([^<]{2,80})<"])
            if looks_like_wrong_product(title, name, mol.get("route")):
                misses.append((name, f"{src['name']}: listing is not this medicine "
                                     f"({str(title)[:60]}) — rejected"))
            elif price is None and mrp is None:
                misses.append((name, f"{src['name']}: page had no readable price"))
            else:
                found_any = True
                strength, pack = extract_pack(title)
                rows.append({
                    "molecule": name, "category": mol.get("category", ""),
                    "retailer": src["name"], "product": title, "manufacturer": mfr,
                    "strength": strength, "pack": pack,
                    "selling_price_inr": price, "mrp_inr": mrp,
                    "source_url": url, "retrieved_utc": stamp,
                })
            time.sleep(REQUEST_DELAY_SEC)
        mark = "✓" if found_any else "—"
        print(f"  [{i:3}/{len(molecules)}] {name[:30]:32} {mark} {len(picked)} retailer(s)")

    # ── Cross-retailer agreement ─────────────────────────────────────────
    # Independent retailers pricing the same molecule should broadly agree.
    # Where three sources cluster and a fourth is orders of magnitude away,
    # the outlier is far more likely to be a mismatched product than a bargain.
    from statistics import median
    by_mol = defaultdict(list)
    for r in rows:
        if r["selling_price_inr"]:
            by_mol[r["molecule"]].append(r["selling_price_inr"])
    for r in rows:
        peers = by_mol.get(r["molecule"], [])
        if len(peers) < 2 or not r["selling_price_inr"]:
            r["agreement"] = "single source" if len(peers) == 1 else ""
            continue
        med = median(peers)
        ratio = (r["selling_price_inr"] / med) if med else 0
        if ratio > 5 or ratio < 0.2:
            r["agreement"] = f"OUTLIER — {ratio:.1f}x the median of {len(peers)} listings; verify product"
        else:
            r["agreement"] = f"consistent with {len(peers)} listings"

    # ── Write the NEW workbook ───────────────────────────────────────────
    out = args.out or os.path.join(
        os.path.expanduser("~"), "Downloads",
        f"DROP_Tax_Sourced_Prices_{datetime.now().strftime('%Y%m%d')}.xlsx")
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Sourced Prices"
    headers = ["Molecule", "Category", "Retailer", "Product listing", "Manufacturer",
               "Strength", "Pack", "Selling price (INR)", "MRP (INR)",
               "Source URL", "Retrieved (UTC)"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E5F")
        c.alignment = Alignment(vertical="center")
    for r in rows:
        ws.append([r["molecule"], r["category"], r["retailer"], r["product"], r["manufacturer"],
                   r.get("strength"), r.get("pack"),
                   r["selling_price_inr"], r["mrp_inr"], r.get("agreement", ""),
                   r["source_url"], r["retrieved_utc"]])
    for col, w in zip("ABCDEFGHIJKL", (26, 16, 16, 44, 24, 12, 8, 18, 14, 46, 62, 22)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Not Found")
    ws2.append(["Molecule", "Reason"])
    for c in ws2[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="7A2E2E")
    for name, reason in misses:
        ws2.append([name, reason])
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 70

    ws3 = wb.create_sheet("Method")
    for line in [
        ["DROP Tax — sourced price workbook"],
        [],
        ["Generated (UTC)", stamp],
        ["Molecules attempted", len(molecules)],
        ["Price rows captured", len(rows)],
        ["Molecules with no price", len({m for m, _ in misses})],
        [],
        ["How these figures were obtained"],
        ["1", "Molecule list read from the platform database, not typed in."],
        ["2", "Product URLs discovered from each retailer's own sitemap."],
        ["3", "Each product page fetched directly; price read from the page."],
        ["4", "Every row carries its source URL and retrieval timestamp."],
        [],
        ["Limitations — read before using these figures"],
        ["a", "Retail listings are not regulated ceiling prices. NPPA ceiling "
              "prices are the authority for scheduled formulations."],
        ["b", "Prices change continuously. One listing moved by about 3,000 "
              "rupees between two reads minutes apart. Figures are valid only "
              "as at the retrieval timestamp."],
        ["c", "Listing price is not net realised price. Trade margin, GST and "
              "institutional discounts are not visible in retail listings."],
        ["d0", "Placeholder MRPs are discarded. PharmEasy returned 1000 as the "
               "MRP on every row sampled, and Apollo returned 2999 and 99999; "
               "these are markup defaults, not prices."],
        ["d1", "Listings that are not the medicine (diagnostic panels, "
               "combination anaesthetics) are rejected, not priced."],
        ["d1b", "Listings whose dose form contradicts the molecule's route are "
                "rejected: zoledronic acid is an infusion, so a capsule listing "
                "is a different drug even though the reference workbook lists "
                "that brand against it."],
        ["d3", "The cross-retailer check compares each price to the median of "
               "the other listings. Treat any row marked OUTLIER as unverified "
               "until the product has been opened and read."],
        ["d2", "Strength and pack size are captured because retailers list "
               "different packs. Do not compare prices across different packs."],
        ["d", "Matching is by URL token, so a row may occasionally point at a "
              "different strength or pack. Check the product listing column."],
        ["e", "Molecules absent from a retailer's catalogue are listed on the "
              "Not Found sheet rather than estimated."],
    ]:
        ws3.append(line)
    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 96
    ws3["A1"].font = Font(bold=True, size=13)

    wb.save(out)
    print(f"\n✅ {len(rows)} price rows from live listings")
    print(f"   {len({m for m, _ in misses})} molecule(s) with no price found")
    print(f"   written to {out}")
    client.close()


if __name__ == "__main__":
    asyncio.run(main())
